"""
data_loader.py
Reads RE_Fraternity_Jun2026_Master.xlsx correctly:
  - Row 3 = header
  - Row 4+ = data (229 records)
  - data_only=True so formulas are resolved to cached values
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from config import SCORE_COLS, REQ_COLS, GAP_COLS, SUMMARY_GROUPS


def load_master_data(path: str) -> pd.DataFrame:
    """
    Load the 'All' sheet and return a clean DataFrame.
    Returns 229 rows × 128 columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["All"]

    # Try to detect header row (some workbooks use row 3 or row 4)
    def _detect_header_row(ws, start=2, end=6):
        for r in range(start, min(end, ws.max_row) + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 200) + 1)]
            vals = [str(v).strip().lower() for v in values if v is not None]
            if any(k in vals for k in ("name", "staff id", "staff_id")):
                return r
        return 3

    header_row_num = _detect_header_row(ws)

    # Build header list from detected header row
    headers = [ws.cell(header_row_num, c).value for c in range(1, ws.max_column + 1)]

    # Read data rows (row after header onwards), stop when Name is empty
    rows = []
    for r in range(header_row_num + 1, ws.max_row + 1):
        name_val = ws.cell(r, 5).value      # col 5 = Name
        if not name_val:
            continue
        row = {}
        for c, h in enumerate(headers, 1):
            if h is not None:
                row[h] = ws.cell(r, c).value
        rows.append(row)

    df = pd.DataFrame(rows)

    # ── Clean types ─────────────────────────────────────────────────────────
    # Numeric: scores, gaps, requirements, summaries, years
    num_cols = (
        SCORE_COLS + REQ_COLS + GAP_COLS +
        [c for grp in SUMMARY_GROUPS.values() for c in grp] +
        ["Age", "Birth Year", "Years in PET", "Years of RE Experience",
         "Years in Salary Grade", "Length in Current Assignment",
         "Age Promoted to Staff or Principal"]
    )
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Dates
    date_cols = ["Chat Date", "Joining Date", "Contract Expire Date",
                 "Last Assesment Date", "Date of Appointment to Current Grade",
                 "Date in Position"]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", format="mixed")

    # String cleanup
    str_cols = ["Name", "Staff ID", "Gender", "Nationality", "Department",
                "Section Name", "Unit Name", "Sub Unit", "Staff Position",
                "SG", "Email Address", "Employment Category", "Chat Status",
                "Assessment Level", "Sub-Disciplines", "Potential", "Strength",
                "Recommendation", "Resource/SME", "Interest", "Preference",
                "Comment/Suggestion", "Assesor1", "Assessor2", "Supervisor", "Remarks"]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace("None", np.nan).replace("nan", np.nan)

    # Ensure Staff ID is string without trailing .0 for numeric IDs
    if "Staff ID" in df.columns:
        def _clean_staff_id(x):
            if pd.isna(x) or x in ("nan", "None"):
                return np.nan
            try:
                return str(int(float(x)))
            except (ValueError, TypeError):
                return str(x).strip()
        df["Staff ID"] = df["Staff ID"].apply(_clean_staff_id)

    df = df.reset_index(drop=True)
    return df


def get_score_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame: Name + all B/K/P/E actual score columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in SCORE_COLS if c in df.columns]
    return df[cols].copy()


def get_gap_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame: Name + all gap columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in GAP_COLS if c in df.columns]
    return df[cols].copy()


def get_req_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame: Name + all requirement/target columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in REQ_COLS if c in df.columns]
    return df[cols].copy()
