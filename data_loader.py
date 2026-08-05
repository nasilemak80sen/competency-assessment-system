"""
data_loader.py – UPDATED
Reads RE_Fraternity_Jul2026_Master.xlsx correctly:
  - Row 3 = header
  - Row 4+ = data (229 records)
  - data_only=True so formulas are resolved to cached values
  
UPDATED: load_ruler_and_tech_mapping() now handles separated ruler structure
where each row has explicit ruler_type + SG (no merging).
"""
import os
from urllib.parse import quote, urljoin
import re
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from config import SCORE_COLS, REQ_COLS, GAP_COLS, SUMMARY_GROUPS, RULER_SHEET, TAB_SEPARATOR_SHEET, CV_LIST_SHEET, CV_LIST_COLUMNS, SHAREPOINT_CV_ROOT_URL, CV_ALLOWED_FILE_TYPES


def _normalize_summary_column_name(name: object) -> str:
    if name is None:
        return ""
    if not isinstance(name, str):
        try:
            name = str(name)
        except Exception:
            return ""
    normalized = name.strip()
    match = re.match(
        r'^(Staff|Principal|Custodian)\s+(Base|Keys|Pacing|Emerging|CTI)\d*$',
        normalized,
        flags=re.IGNORECASE,
    )
    if match:
        prefix = match.group(1).title()
        suffix = match.group(2).upper() if match.group(2).upper() == 'CTI' else match.group(2).title()
        return f"{prefix} {suffix}"
    return normalized


def _safe_rule_val(val):
    if val is None:
        return None

    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)

    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def load_master_data(path: str) -> pd.DataFrame:
    """
    Load the 'All' sheet and return a clean DataFrame.
    Returns 229 rows × 128 columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["All"]

    # Build and normalize header list from row 3
    headers = [
        _normalize_summary_column_name(ws.cell(3, c).value)
        for c in range(1, ws.max_column + 1)
    ]

    # Read data rows (row 4 onwards), stop when Name is empty
    rows = []
    for r in range(4, ws.max_row + 1):
        name_val = ws.cell(r, 5).value      # col 5 = Name
        if not name_val:
            continue
        row = {}
        for c, h in enumerate(headers, 1):
            if h is not None:
                row[h] = ws.cell(r, c).value
        rows.append(row)

    df = pd.DataFrame(rows)

    # Normalize malformed summary header names such as Staff Base2 -> Staff Base
    df.rename(
        columns={col: _normalize_summary_column_name(col) for col in df.columns},
        inplace=True,
    )

    # ── Clean types ─────────────────────────────────────────────────────────
    # Numeric: scores, gaps, requirements, summaries, years
    num_cols = (
        SCORE_COLS + REQ_COLS + GAP_COLS +
        [c for grp in SUMMARY_GROUPS.values() for c in grp] +
        ["Age", "Birth Year", "Years in PET", "Years of RE Experience",
         "Years in Salary Grade", "Length in Current Assignment",
         "Age Promoted to Staff or Principal"]
    )
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Dates
    date_cols = ["Chat Date", "Joining Date", "Contract Expire Date",
                 "Last Assesment Date", "Date of Appointment to Current Grade",
                 "Date in Position"]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", format="mixed")

    # String cleanup
    str_cols = ["Name", "Staff ID", "Gender", "Nationality", "Department",
                "Section Name", "Unit Name", "Sub Unit", "Staff Position",
                "SG", "Ruler Type", "Background", "Email Address", "Employment Category", "Chat Status",
                "Assessment Level", "Sub-Disciplines", "Potential", "Strength",
                "Recommendation", "Resource/SME", "Interest", "Preference",
                "Comment/Suggestion", "Assesor1", "Assessor2", "Supervisor", "Remarks"]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace("None", np.nan).replace("nan", np.nan)

    # Ensure Staff ID is string without trailing .0 for numeric IDs
    if "Staff ID" in df.columns:
        def _clean_staff_id(x):
            if pd.isna(x) or x in ("nan", "None"):
                return np.nan
            try:
                return str(int(float(x)))
            except (ValueError, TypeError):
                return str(x).strip()
        df["Staff ID"] = df["Staff ID"].apply(_clean_staff_id)

    df = df.reset_index(drop=True)
    return df


def load_ruler_and_tech_mapping(
    path: str,
    verbose: bool = False,
) -> tuple:
    """
    Load ruler requirements and competency labels.

    Expected Ruler sheet structure:
        Row 4:
            Column A = Ruler Type
            Column B = SG
            Column C = Position
            Column D onward = competency requirements

        Row 5 onward:
            Base / RDP / RMS / RSS
            P1 / P2 / ... / P10
    """
    wb = openpyxl.load_workbook(
        path,
        data_only=True,
    )

    ruler_map = {}
    tech_labels = {}

    # -------------------------------------------------------------------------
    # TAB SEPARATOR
    # -------------------------------------------------------------------------

    if TAB_SEPARATOR_SHEET in wb.sheetnames:
        ws_labels = wb[TAB_SEPARATOR_SHEET]

        for row_number in range(
            1,
            ws_labels.max_row + 1,
        ):
            row_values = [
                ws_labels.cell(
                    row_number,
                    column_number,
                ).value
                for column_number in range(
                    1,
                    ws_labels.max_column + 1,
                )
            ]

            for index, raw_code in enumerate(
                row_values
            ):
                if raw_code is None:
                    continue

                code = str(raw_code).strip()

                if (
                    code in SCORE_COLS
                    and index + 1 < len(row_values)
                ):
                    full_name = row_values[index + 1]

                    if full_name is not None:
                        tech_labels[code] = str(
                            full_name
                        ).strip()

    # -------------------------------------------------------------------------
    # RULER SHEET VALIDATION
    # -------------------------------------------------------------------------

    if RULER_SHEET not in wb.sheetnames:
        raise FileNotFoundError(
            f"Sheet '{RULER_SHEET}' was not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[RULER_SHEET]

    # Row 4 contains the headers.
    header_row = [
        ws.cell(4, column_number).value
        for column_number in range(
            1,
            ws.max_column + 1,
        )
    ]

    # Clean header strings.
    cleaned_headers = [
        str(header).strip()
        if header is not None
        else None
        for header in header_row
    ]

    missing_competencies = [
        competency
        for competency in SCORE_COLS
        if competency not in cleaned_headers
    ]

    if missing_competencies:
        raise ValueError(
            "Ruler sheet row 4 is missing these "
            f"competency headers: {missing_competencies}"
        )

    # Identify competency columns from the header row.
    score_column_indices = [
        (column_number, header)
        for column_number, header in enumerate(
            cleaned_headers,
            start=1,
        )
        if header in SCORE_COLS
    ]

    if verbose:
        print(
            "Ruler competency columns detected:",
            len(score_column_indices),
        )

    # -------------------------------------------------------------------------
    # LOAD RULER ROWS
    # -------------------------------------------------------------------------

    row_count = 0

    for row_number in range(
        5,
        ws.max_row + 1,
    ):
        # IMPORTANT:
        # Workbook Column A = Ruler Type
        # Workbook Column B = SG
        # Workbook Column C = Position
        raw_ruler_type = ws.cell(
            row_number,
            1,
        ).value

        raw_sg = ws.cell(
            row_number,
            2,
        ).value

        # Skip fully empty rows.
        if raw_ruler_type is None and raw_sg is None:
            continue

        # Skip rows without a valid SG.
        if raw_sg is None:
            continue

        sg = str(raw_sg).strip().upper()

        # Only accept P1 through P10.
        if not re.fullmatch(
            r"P(?:[1-9]|10)",
            sg,
        ):
            if verbose:
                print(
                    f"Skipping row {row_number}: "
                    f"invalid SG={raw_sg!r}"
                )

            continue

        # Normalize ruler type.
        if (
            raw_ruler_type is None
            or str(raw_ruler_type).strip() == ""
        ):
            ruler_type = "BASE"
        else:
            ruler_type = str(
                raw_ruler_type
            ).strip().upper()

        aliases = {
            "BASE": "BASE",
            "NO RULER ASSIGNED": "BASE",
            "RDP": "RDP",
            "RMS": "RMS",
            "RSS": "RSS",
        }

        ruler_type = aliases.get(
            ruler_type,
            ruler_type,
        )

        # Reject unrelated ruler types.
        if ruler_type not in {
            "BASE",
            "RDP",
            "RMS",
            "RSS",
        }:
            if verbose:
                print(
                    f"Skipping row {row_number}: "
                    f"invalid ruler={ruler_type!r}"
                )

            continue

        row_data = {}

        for column_number, competency in (
            score_column_indices
        ):
            raw_score = ws.cell(
                row_number,
                column_number,
            ).value

            score = _safe_rule_val(
                raw_score
            )

            # Preserve zero as an explicit requirement.
            if score is not None:
                row_data[competency] = score

        if not row_data:
            if verbose:
                print(
                    f"Skipping row {row_number}: "
                    "no numeric competency requirements"
                )

            continue

        ruler_map.setdefault(
            ruler_type,
            {},
        )[sg] = row_data

        row_count += 1

    # -------------------------------------------------------------------------
    # FINAL VALIDATION
    # -------------------------------------------------------------------------

    if not ruler_map:
        raise ValueError(
            "No ruler data was loaded from the Ruler sheet."
        )

    career_rulers = {
        "RDP",
        "RMS",
        "RSS",
    }

    detected_rulers = set(
        ruler_map.keys()
    )

    missing_career_rulers = (
        career_rulers - detected_rulers
    )

    if missing_career_rulers:
        raise ValueError(
            "Ruler sheet is missing the following "
            "progression ruler types: "
            f"{sorted(missing_career_rulers)}. "
            f"Detected ruler types: "
            f"{sorted(detected_rulers)}"
        )

    if verbose:
        print("\nRuler loading completed")
        print("Rows loaded:", row_count)
        print(
            "Ruler types:",
            sorted(ruler_map.keys()),
        )

        for ruler_type in sorted(
            ruler_map.keys()
        ):
            grades = sorted(
                ruler_map[ruler_type].keys(),
                key=lambda value: int(value[1:]),
            )

            print(
                f"  {ruler_type}: {grades}"
            )

    return ruler_map, tech_labels
    """
    Load the Ruler and Tab Separator sheets from the Excel master workbook.
    
    NEW STRUCTURE (separated rows):
    - Row 4: Headers (B1, B2, ... K1, ... P1, P2, P3, P4, P5, E1, E2)
    - Row 5+: Each row has explicit Ruler Type (Col B) + SG (Col C)
      - Ruler Type can be blank (→ "BASE") or "RDP", "RMS", "RSS"
      - SG is always present (P1, P2, P3, ... P10)
    
    Returns:
        (ruler_map, tech_labels)
        ruler_map = {ruler_type: {sg: {competency: required_level}}}
        tech_labels = {competency_code: full_name}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ruler_map = {}
    tech_labels = {}

    # ─────── Load competency full names from Tab Separator ─────────────────
    if TAB_SEPARATOR_SHEET in wb.sheetnames:
        ws = wb[TAB_SEPARATOR_SHEET]
        for r in range(1, ws.max_row + 1):
            row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for idx, code in enumerate(row_values):
                if code is None:
                    continue
                code_str = str(code).strip()
                if code_str in SCORE_COLS and idx + 1 < len(row_values):
                    full_name = row_values[idx + 1]
                    if full_name is not None:
                        tech_labels[code_str] = str(full_name).strip()

    # ─────── Load Ruler Sheet ────────────────────────────────────────────────
    if RULER_SHEET not in wb.sheetnames:
        raise FileNotFoundError(f"Sheet '{RULER_SHEET}' not found in workbook")

    ws = wb[RULER_SHEET]
    
    # Read header row (row 4)
    header_row = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    
    # Validate all expected competency columns are present
    header_set = set(h for h in header_row if h is not None)
    missing_cols = set(SCORE_COLS) - header_set
    if missing_cols:
        raise ValueError(
            f"Ruler sheet header (row 4) missing competencies: {sorted(missing_cols)}"
        )
    
    # Build index of score column positions
    score_col_indices = [(c, h) for c, h in enumerate(header_row, 1) if h in SCORE_COLS]

    # Parse data rows (starting from row 5)
    row_count = 0
    for r in range(5, ws.max_row + 1):
        ruler_type_val = ws.cell(r, 1).value  # Column B = Ruler Type
        sg_val = ws.cell(r, 2).value          # Column C = SG (Grade)

        print(f"Row {r}: Ruler Type={ruler_type_val}, SG={sg_val}")  # Debugging output
        
        # Skip rows without a grade
        if not sg_val:
            continue
        
        sg = str(sg_val).strip()
        
        # Determine ruler type: blank → "BASE", otherwise use value
        if not ruler_type_val or str(ruler_type_val).strip() == "":
            ruler_type = "BASE"
        else:
            ruler_type = str(ruler_type_val).strip()
        
        # Extract competency scores for this ruler type + grade combination
        row_data = {}
        for col_pos, (c, comp_code) in enumerate(score_col_indices):
            cell_value = ws.cell(r, c).value
            score = _safe_rule_val(cell_value)

            print(f"  Competency {comp_code}: Raw={cell_value}, Parsed={score}")  # Debugging output    
            
            # Only store non-None scores
            if score is not None:
                row_data[comp_code] = score
        
        # Only store this row if it has at least one competency score
        if row_data:
            ruler_map.setdefault(ruler_type, {})[sg] = row_data
            row_count += 1

    if not ruler_map:
        raise ValueError("No ruler data loaded from sheet")

    return ruler_map, tech_labels


def get_score_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame: Name + all B/K/P/E actual score columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in SCORE_COLS if c in df.columns]
    return df[cols].copy()


def get_gap_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Return sub-DataFrame: Name + all gap columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in GAP_COLS if c in df.columns]
    return df[cols].copy()


def get_req_matrix(df: pd.DataFrame) -> pd.DataFrame:

    """Return sub-DataFrame: Name + all requirement/target columns."""
    cols = ["Name", "Staff Position", "SG", "Department"] + \
           [c for c in REQ_COLS if c in df.columns]
    return df[cols].copy()

def _clean_cv_value(value):
    """
    Convert blank Excel values into None and trim text values.
    """
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, str):
        value = value.strip()

        if not value:
            return None

    return value


def _normalise_staff_id(value):
    """
    Normalize Staff IDs read from Excel.

    Examples:
        1019423.0 -> 1019423
        " 1019423 " -> 1019423
    """
    value = _clean_cv_value(value)

    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    cleaned = str(value).strip()

    cleaned = re.sub(
        r"\.0$",
        "",
        cleaned,
    )

    return cleaned or None


def _normalise_person_name(value):
    """
    Normalize names for matching only.

    The original name is still retained for display.
    """
    value = _clean_cv_value(value)

    if value is None:
        return None

    normalized = str(value).lower().strip()

    replacements = {
        "@": " ",
        ".": " ",
        ",": " ",
        "'": "",
        "-": " ",
        "_": " ",
        "(": " ",
        ")": " ",
        "&": " and ",
    }

    for old_value, new_value in replacements.items():
        normalized = normalized.replace(
            old_value,
            new_value,
        )

    normalized = re.sub(
        r"\s+",
        " ",
        normalized,
    )

    return normalized.strip() or None


def _build_sharepoint_cv_url(hyperlink_target):
    """
    Convert the CV-list hyperlink target into a complete
    SharePoint HTTPS URL.

    The current workbook contains relative targets such as:

        Executive/Employee%20Name/CV.pdf
    """
    target = _clean_cv_value(
        hyperlink_target
    )

    if target is None:
        return None

    target = str(target).strip()

    # Already a complete web URL.
    if target.lower().startswith(
        ("https://", "http://")
    ):
        return target

    # Do not expose file:/// links as SharePoint URLs.
    if target.lower().startswith(
        "file:///"
    ):
        return None

    # Convert Windows separators in case the target uses them.
    target = target.replace("\\", "/")

    # Preserve existing %20 encoding while escaping literal spaces.
    target = target.replace(" ", "%20")

    return urljoin(
        SHAREPOINT_CV_ROOT_URL,
        target.lstrip("/"),
    )


def load_cv_list(path: str) -> pd.DataFrame:
    """
    Load the CV list worksheet, including actual Excel hyperlink targets.

    pandas.read_excel() reads only the visible hyperlink text,
    such as 'Open in SharePoint'. openpyxl is required to read
    cell.hyperlink.target.
    """
    workbook = openpyxl.load_workbook(
        path,
        data_only=False,
        read_only=False,
    )

    if CV_LIST_SHEET not in workbook.sheetnames:
        return pd.DataFrame(
            columns=CV_LIST_COLUMNS
        )

    worksheet = workbook[CV_LIST_SHEET]

    headers = {}

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        raw_header = worksheet.cell(
            row=1,
            column=column_number,
        ).value

        if raw_header is None:
            continue

        cleaned_header = str(
            raw_header
        ).strip()

        headers[cleaned_header] = (
            column_number
        )

    records = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        record = {}

        for column_name in CV_LIST_COLUMNS:
            column_number = headers.get(
                column_name
            )

            if column_number is None:
                continue

            record[column_name] = (
                _clean_cv_value(
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                )
            )

        # Skip completely blank rows.
        if not any(
            value is not None
            for value in record.values()
        ):
            continue

        # Extract actual hyperlink targets.
        local_link_column = headers.get(
            "Local File Link"
        )

        sharepoint_link_column = headers.get(
            "SharePoint URL"
        )

        local_target = None
        sharepoint_target = None

        if local_link_column is not None:
            local_cell = worksheet.cell(
                row=row_number,
                column=local_link_column,
            )

            if local_cell.hyperlink:
                local_target = (
                    local_cell.hyperlink.target
                )

        if sharepoint_link_column is not None:
            sharepoint_cell = worksheet.cell(
                row=row_number,
                column=sharepoint_link_column,
            )

            if sharepoint_cell.hyperlink:
                sharepoint_target = (
                    sharepoint_cell.hyperlink.target
                )

        record["Local File Link"] = (
            local_target
            or record.get("Local File Path")
        )

        record["SharePoint URL"] = (
            _build_sharepoint_cv_url(
                sharepoint_target
            )
        )

        record["Staff ID"] = (
            _normalise_staff_id(
                record.get("Staff ID")
            )
        )

        record["Normalized Name"] = (
            _normalise_person_name(
                record.get("Name")
            )
        )

        file_type = _clean_cv_value(
            record.get("File Type")
        )

        if file_type is not None:
            record["File Type"] = (
                str(file_type)
                .strip()
                .upper()
            )

        modified_value = record.get("Modified Date")
        modified_date = None
        if modified_value is not None and not pd.isna(modified_value):
            modified_date = pd.to_datetime(
                modified_value,
                errors="coerce",
            )

        record["Modified Date"] = (
            None
            if modified_date is None or pd.isna(modified_date)
            else modified_date.to_pydatetime()
        )

        records.append(record)

    cv_dataframe = pd.DataFrame(
        records
    )

    if cv_dataframe.empty:
        return cv_dataframe

    # Keep only supported document types.
    cv_dataframe = cv_dataframe[
        cv_dataframe["File Type"].isin(
            CV_ALLOWED_FILE_TYPES
        )
    ].copy()

    # A document record must have a file name.
    cv_dataframe = cv_dataframe[
        cv_dataframe["CV File Name"]
        .notna()
    ].copy()

    return cv_dataframe.reset_index(
        drop=True
    )