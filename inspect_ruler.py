import openpyxl
from pathlib import Path
path = Path(r'RE_Fraternity_Jun2026_Master.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS', wb.sheetnames)
ws = wb['Ruler']
print('max_row', ws.max_row, 'max_col', ws.max_column)
for r in range(1, min(ws.max_row, 20)+1):
    vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20)+1)]
    print(r, vals)
