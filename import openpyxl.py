import openpyxl
from pathlib import Path

path = Path(r"C:\Users\mnabielizzuddin.radz\OneDrive - PETRONAS\Reservoir Engineering\Programming_Python_Projects\Competency Assessment System\RE Fraternity Jul2026_Master.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Ruler']

# Check row 4 (headers)
header_row = [ws.cell(4, c).value for c in range(1, 30)]
print("Row 4 (Headers):", header_row[:15])

# Check what grades exist
for r in range(5, 20):
    sg = ws.cell(r, 3).value
    ruler_type = ws.cell(r, 2).value
    if sg:
        print(f"Row {r}: Ruler Type={ruler_type}, SG={sg}")